import sys   # لإدارة مسارات Python وإنهاء البرنامج
import os    # للتعامل مع الملفات والمسارات ومتغيرات البيئة
from urllib.parse import urlparse, unquote

# ── إضافة مسار المكتبات المحلية تلقائياً عند تشغيل الملف مباشرةً ──────────
_base_dir = os.path.dirname(os.path.abspath(__file__))          # مسار المجلد الذي يوجد فيه app.py
_local_sp = os.path.join(_base_dir, 'python', 'Lib', 'site-packages')  # مسار المكتبات المحلية المرفقة مع المشروع
if os.path.isdir(_local_sp) and _local_sp not in sys.path:      # إذا كان المجلد موجوداً ولم يُضَف بعد
    sys.path.insert(0, _local_sp)                               # أضفه في أول قائمة مسارات Python

try:
    from flask import Flask, render_template, request, jsonify, redirect, url_for, session, flash, g, send_from_directory
    # ↑ استيراد أدوات Flask الأساسية: الإطار - تصيير HTML - الطلب - JSON - إعادة التوجيه - الرابط - الجلسة - الفلاش - السياق - إرسال ملفات
    from werkzeug.security import generate_password_hash, check_password_hash  # لتشفير كلمات المرور والتحقق منها
    from functools import wraps      # للحفاظ على اسم الدالة عند استخدام المُزخرفات (decorators)
    from dotenv import load_dotenv   # لتحميل متغيرات البيئة من ملف .env
except ImportError as e:
    print(f'\n[ERROR] مكتبة ناقصة: {e}')  # طباعة اسم المكتبة الناقصة
    print('       شغّل التطبيق من run.bat أو ثبّت المكتبات: pip install -r requirements.txt')
    input('\nاضغط Enter للإغلاق ...')      # توقف حتى يقرأ المستخدم الرسالة
    sys.exit(1)                             # إنهاء البرنامج برمز خطأ 1

import sqlite3              # للتعامل مع قاعدة بيانات SQLite المدمجة في Python
import io                   # للتعامل مع البيانات في الذاكرة كملفات وهمية (للـ PDF والـ Excel)
import datetime             # للحصول على التاريخ والوقت الحالي
import time                 # للحصول على الوقت بالثواني (مستخدم في Rate Limiting)
import shutil               # لنسخ ونقل الملفات (للنسخ الاحتياطية)
from collections import defaultdict  # قاموس يعطي قيمة افتراضية إذا لم يوجد المفتاح (مستخدم في Rate Limiting)

load_dotenv()  # تحميل متغيرات البيئة من ملف .env الموجود في نفس المجلد

app = Flask(__name__)  # إنشاء تطبيق Flask وتحديد اسمه من اسم الملف الحالي
app.secret_key = os.environ.get('SECRET_KEY', 'green-economy-secret-key-2024')
# ↑ المفتاح السري لتشفير الجلسات – يُقرأ من .env وإلا يستخدم القيمة الافتراضية

DB_PATH = os.path.join(os.path.dirname(__file__), 'database', 'green_economy.db')  # المسار الكامل لملف قاعدة البيانات
CARBON_ALERT_THRESHOLD = float(os.environ.get('CARBON_ALERT_THRESHOLD', '1000'))  # الحد الافتراضي للتنبيه على انبعاثات الكربون (1000 طن)

# ── إعدادات البريد الإلكتروني ──────────────────────────────────────────────
MAIL_ENABLED  = os.environ.get('MAIL_ENABLED', 'false').lower() == 'true'  # هل إرسال البريد مفعّل؟ (true/false)
MAIL_SERVER   = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')            # عنوان خادم البريد الصادر
MAIL_PORT     = int(os.environ.get('MAIL_PORT', '587'))                    # منفذ خادم البريد (587 للـ TLS)
MAIL_USERNAME = os.environ.get('MAIL_USERNAME', '')                        # البريد الإلكتروني المُرسِل
MAIL_PASSWORD = os.environ.get('MAIL_PASSWORD', '')                        # كلمة مرور البريد المُرسِل
MAIL_TO       = os.environ.get('MAIL_TO', '')                              # البريد الإلكتروني المُستقبِل للتنبيهات

# ── قاعدة البيانات: SQLite أو PostgreSQL أو MySQL/MariaDB ─────────────────
DATABASE_URL = os.environ.get('DATABASE_URL', '').strip()
_db_scheme = ''
if DATABASE_URL:
    _db_scheme = urlparse(DATABASE_URL).scheme.split('+', 1)[0].lower()
USE_POSTGRES = _db_scheme in ('postgres', 'postgresql')
USE_MYSQL = _db_scheme in ('mysql', 'mariadb')
USE_SQLITE = not (USE_POSTGRES or USE_MYSQL)
DB_TYPE_LABEL = 'PostgreSQL' if USE_POSTGRES else 'MySQL/MariaDB' if USE_MYSQL else 'SQLite'
SETTINGS_KEY_COL = 'setting_key' if USE_MYSQL else 'key'


# ══════════════════════════════════════════════════════════════════════════════
# اتصال قاعدة البيانات (SQLite / PostgreSQL / MySQL)
# ══════════════════════════════════════════════════════════════════════════════

if USE_POSTGRES:  # إذا كان المشروع مضبوطاً على PostgreSQL
    import psycopg2              # مكتبة الاتصال بالـ PostgreSQL
    import psycopg2.extras       # أدوات إضافية مثل RealDictCursor (لإرجاع الصفوف كقواميس)

    def get_db():
        """ يُرجع اتصال PostgreSQL الخاص بالطلب الحالي """
        if 'db' not in g:  # إذا لم يكن هناك اتصال مفتوح في هذا الطلب
            g.db = psycopg2.connect(DATABASE_URL)  # افتح اتصال جديد
            g.db.autocommit = False  # عطّل autocommit حتى نتحكم يدوياً في الحفظ
        return g.db  # أرجع الاتصال المخزّن في g (سياق الطلب)

    @app.teardown_appcontext          # تُنفّذ تلقائياً بعد انتهاء كل طلب HTTP
    def close_db(exc=None):
        db = g.pop('db', None)  # اسحب الاتصال من سياق الطلب
        if db:  # إذا كان مفتوحاً
            db.close()  # أغلقه لتحرير الموارد

    def query(sql, params=(), one=False):
        """ تنفيذ استعلام SELECT وإرجاع النتائج """
        sql = sql.replace('?', '%s').replace("datetime('now')", 'NOW()')  # تحويل صيغة SQLite إلى PostgreSQL
        cur = get_db().cursor(cursor_factory=psycopg2.extras.RealDictCursor)  # cursor يُرجع كل صف كقاموس
        cur.execute(sql, params)  # تنفيذ الاستعلام
        rows = cur.fetchall()     # جلب كل الصفوف
        return (rows[0] if rows else None) if one else rows  # إذا one=True أرجع صفاً واحداً فقط

    def execute(sql, params=()):
        """ تنفيذ استعلام INSERT/UPDATE/DELETE """
        sql = sql.replace('?', '%s').replace("datetime('now')", 'NOW()')  # تحويل الصيغة
        db = get_db()        # اجلب اتصال قاعدة البيانات
        cur = db.cursor()    # افتح cursor
        cur.execute(sql, params)  # تنفيذ الاستعلام
        db.commit()          # حفظ التغييرات في قاعدة البيانات
        return cur           # أرجع الـ cursor (المستدعي يقدر يتحقق من lastrowid)

    def init_db():
        """ إنشاء الجداول في PostgreSQL إذا لم تكن موجودة """
        db = psycopg2.connect(DATABASE_URL)  # اتصال مباشر لإنشاء الجداول
        cur = db.cursor()    # cursor لتنفيذ الاستعلامات
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id            SERIAL PRIMARY KEY,
                username      TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                role          TEXT NOT NULL DEFAULT 'viewer',
                created_at    TIMESTAMP DEFAULT NOW()
            );
            CREATE TABLE IF NOT EXISTS institutions (
                id                          SERIAL PRIMARY KEY,
                name                        TEXT    NOT NULL,
                year                        INTEGER NOT NULL,
                energy_consumption          REAL    NOT NULL,
                renewable_energy_percentage REAL    NOT NULL,
                carbon_emissions            REAL    NOT NULL,
                green_projects              INTEGER NOT NULL DEFAULT 0,
                water_usage                 REAL    NOT NULL DEFAULT 0,
                waste_recycling_percentage  REAL    NOT NULL DEFAULT 0,
                created_at  TIMESTAMP DEFAULT NOW(),
                updated_at  TIMESTAMP DEFAULT NOW()
            );
            CREATE TABLE IF NOT EXISTS settings (
                key   TEXT PRIMARY KEY,
                value TEXT NOT NULL DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS audit_log (
                id         SERIAL PRIMARY KEY,
                username   TEXT    NOT NULL,
                action     TEXT    NOT NULL,
                details    TEXT    DEFAULT '',
                created_at TIMESTAMP DEFAULT NOW()
            );
        """)
        db.commit()   # حفظ إنشاء الجداول
        db.close()    # أغلق الاتصال بعد الإنشاء

elif USE_MYSQL:  # إذا كان المشروع مضبوطاً على MySQL/MariaDB
    import pymysql
    import pymysql.cursors

    _MYSQL_CFG = None

    def _parse_mysql_url(url):
        parsed = urlparse(url)
        db_name = (parsed.path or '').lstrip('/')
        if not db_name:
            raise ValueError('DATABASE_URL يجب أن يحتوي على اسم قاعدة البيانات')
        return {
            'host': parsed.hostname or 'localhost',
            'user': unquote(parsed.username) if parsed.username else 'root',
            'password': unquote(parsed.password) if parsed.password else '',
            'database': db_name,
            'port': parsed.port or 3306,
            'charset': 'utf8mb4',
            'cursorclass': pymysql.cursors.DictCursor,
        }

    def _get_mysql_config():
        global _MYSQL_CFG
        if _MYSQL_CFG is None:
            _MYSQL_CFG = _parse_mysql_url(DATABASE_URL)
        return dict(_MYSQL_CFG)

    def _ensure_mysql_database():
        cfg = _get_mysql_config()
        db_name = cfg.get('database')
        server_cfg = {k: v for k, v in cfg.items() if k != 'database'}
        db = pymysql.connect(**server_cfg)
        cur = db.cursor()
        cur.execute(
            f"CREATE DATABASE IF NOT EXISTS `{db_name}` "
            "CHARACTER SET utf8mb4 COLLATE utf8mb4_general_ci"
        )
        db.commit()
        db.close()

    def get_db():
        """ يُرجع اتصال MySQL الخاص بالطلب الحالي """
        if 'db' not in g:
            cfg = _get_mysql_config()
            g.db = pymysql.connect(**cfg)
            g.db.autocommit(False)
        return g.db

    @app.teardown_appcontext          # تُنفّذ تلقائياً بعد انتهاء كل طلب HTTP
    def close_db(exc=None):
        db = g.pop('db', None)  # اسحب الاتصال من سياق الطلب
        if db:  # إذا كان مفتوحاً
            db.close()  # أغلقه لتحرير الموارد

    def query(sql, params=(), one=False):
        """ تنفيذ استعلام SELECT وإرجاع النتائج """
        sql = sql.replace('?', '%s').replace("datetime('now')", 'NOW()')
        cur = get_db().cursor()
        cur.execute(sql, params)
        rows = cur.fetchall()
        return (rows[0] if rows else None) if one else rows

    def execute(sql, params=()):
        """ تنفيذ استعلام INSERT/UPDATE/DELETE """
        sql = sql.replace('?', '%s').replace("datetime('now')", 'NOW()')
        db = get_db()
        cur = db.cursor()
        cur.execute(sql, params)
        db.commit()
        return cur

    def init_db():
        """ إنشاء الجداول في MySQL إذا لم تكن موجودة """
        _ensure_mysql_database()
        cfg = _get_mysql_config()
        db = pymysql.connect(**cfg)
        cur = db.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id            INT AUTO_INCREMENT PRIMARY KEY,
                username      VARCHAR(150) NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                role          VARCHAR(20) NOT NULL DEFAULT 'viewer',
                created_at    TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS institutions (
                id                          INT AUTO_INCREMENT PRIMARY KEY,
                name                        VARCHAR(255) NOT NULL,
                year                        INT NOT NULL,
                energy_consumption          DOUBLE NOT NULL,
                renewable_energy_percentage DOUBLE NOT NULL,
                carbon_emissions            DOUBLE NOT NULL,
                green_projects              INT NOT NULL DEFAULT 0,
                water_usage                 DOUBLE NOT NULL DEFAULT 0,
                waste_recycling_percentage  DOUBLE NOT NULL DEFAULT 0,
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                setting_key VARCHAR(64) PRIMARY KEY,
                value TEXT NOT NULL
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS audit_log (
                id         INT AUTO_INCREMENT PRIMARY KEY,
                username   VARCHAR(150) NOT NULL,
                action     TEXT NOT NULL,
                details    TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
        """)
        db.commit()
        db.close()

else:  # وضع SQLite الافتراضي للتشغيل المحلي
    def get_db():
        """ يُرجع اتصال SQLite الخاص بالطلب الحالي """
        if 'db' not in g:  # إذا لم يكن هناك اتصال مفتوح
            g.db = sqlite3.connect(DB_PATH)  # افتح ملف قاعدة البيانات
            g.db.row_factory = sqlite3.Row   # جعل كل صف قابلاً للوصول بالاسم row['column']
            g.db.execute('PRAGMA journal_mode=WAL')  # تفعيل WAL لأداء أفضل عند الكتابة المتزامنة
            g.db.execute('PRAGMA foreign_keys=ON')   # تفعيل قيود المفاتيح الخارجية
        return g.db  # أرجع الاتصال

    @app.teardown_appcontext          # تُنفّذ تلقائياً بعد انتهاء كل طلب
    def close_db(exc=None):
        db = g.pop('db', None)  # اسحب الاتصال من سياق الطلب
        if db:                  # إذا كان مفتوحاً
            db.close()          # أغلقه لتحرير الموارد

    def query(sql, params=(), one=False):
        """ تنفيذ استعلام SELECT وإرجاع النتائج """
        cur = get_db().execute(sql, params)  # تنفيذ استعلام SQL
        rows = cur.fetchall()               # جلب كل الصفوف
        return (rows[0] if rows else None) if one else rows  # إذا one=True أرجع صفاً واحداً فقط

    def execute(sql, params=()):
        """ تنفيذ استعلام INSERT/UPDATE/DELETE """
        db = get_db()        # اجلب اتصال قاعدة البيانات
        cur = db.execute(sql, params)  # تنفيذ الاستعلام
        db.commit()          # حفظ التغييرات
        return cur           # أرجع الـ cursor

    def init_db():
        """ إنشاء ملف SQLite وجداوله إذا لم تكن موجودة """
        os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)  # أنشئ مجلد database/ إن لم يكن موجوداً
        db = sqlite3.connect(DB_PATH)  # افتح/أنشئ ملف قاعدة البيانات
        db.row_factory = sqlite3.Row   # جعل الصفوف قابلة للوصول بااسم العمود
        db.executescript("""
            CREATE TABLE IF NOT EXISTS users (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                username      TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                role          TEXT NOT NULL DEFAULT 'viewer',
                created_at    TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS institutions (
                id                          INTEGER PRIMARY KEY AUTOINCREMENT,
                name                        TEXT    NOT NULL,
                year                        INTEGER NOT NULL,
                energy_consumption          REAL    NOT NULL,
                renewable_energy_percentage REAL    NOT NULL,
                carbon_emissions            REAL    NOT NULL,
                green_projects              INTEGER NOT NULL DEFAULT 0,
                water_usage                 REAL    NOT NULL DEFAULT 0,
                waste_recycling_percentage  REAL    NOT NULL DEFAULT 0,
                created_at  TEXT DEFAULT (datetime('now')),
                updated_at  TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS settings (
                key   TEXT PRIMARY KEY,
                value TEXT NOT NULL DEFAULT ''
            );
            CREATE TABLE IF NOT EXISTS audit_log (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                username   TEXT    NOT NULL,
                action     TEXT    NOT NULL,
                details    TEXT    DEFAULT '',
                created_at TEXT    DEFAULT (datetime('now'))
            );
        """)
        db.commit()   # حفظ إنشاء الجداول
        db.close()    # أغلق الاتصال بعد الإنشاء

        # ── Migrations: إضافة أعمدة قد تكون ناقصة في قواعد بيانات قديمة ──
        db = sqlite3.connect(DB_PATH)  # افتح اتصال جديد لتنفيذ الـ Migrations
        existing_cols = {row[1] for row in db.execute('PRAGMA table_info(users)').fetchall()}  # اجلب أسماء الأعمدة الحالية
        if 'role' not in existing_cols:  # إذا لم يكن عمود role موجوداً
            db.execute("ALTER TABLE users ADD COLUMN role TEXT NOT NULL DEFAULT 'viewer'")  # أضف عمود الدور
            db.execute("UPDATE users SET role='admin' WHERE id=(SELECT MIN(id) FROM users)")  # اجعل أول مستخدم مسؤولاً
            db.commit()  # حفظ
        existing_cols = {row[1] for row in db.execute('PRAGMA table_info(institutions)').fetchall()}  # أعمدة جدول institutions
        if 'water_usage' not in existing_cols:  # إذا لم يوجد عمود استهلاك المياه
            db.execute('ALTER TABLE institutions ADD COLUMN water_usage REAL NOT NULL DEFAULT 0')  # أضف العمود
            db.commit()  # حفظ
        if 'waste_recycling_percentage' not in existing_cols:  # إذا لم يوجد عمود نسبة إعادة التدوير
            db.execute('ALTER TABLE institutions ADD COLUMN waste_recycling_percentage REAL NOT NULL DEFAULT 0')  # أضف العمود
            db.commit()  # حفظ
        db.close()  # أغلق الاتصال بعد المـ Migrations


# ── إدارة إعدادات النظام ───────────────────────────────────────────────────
def upsert_setting(key, value):
    """ حفظ إعداد مع دعم جميع أنواع قواعد البيانات """
    if USE_POSTGRES:
        sql = 'INSERT INTO settings (key, value) VALUES (?, ?) ON CONFLICT (key) DO UPDATE SET value=EXCLUDED.value'
    elif USE_MYSQL:
        sql = f'INSERT INTO settings ({SETTINGS_KEY_COL}, value) VALUES (?, ?) ON DUPLICATE KEY UPDATE value=VALUES(value)'
    else:
        sql = 'INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)'
    execute(sql, (key, value))


# ── حد الكربون الديناميكي وسجل التدقيق ─────────────────────────────────────
def get_carbon_threshold():
    """يقرأ حد انبعاثات الكربون من إعدادات قاعدة البيانات"""
    try:
        row = query(f'SELECT value FROM settings WHERE {SETTINGS_KEY_COL}=?', ('carbon_threshold',), one=True)  # ابحث عن الحد في جدول settings
        if row and row['value']:       # إذا وُجدت قيمة
            return float(row['value'])  # أرجعها كرقم
    except Exception:
        pass  # تجاهل أي خطأ
    return CARBON_ALERT_THRESHOLD  # أرجع القيمة الافتراضية


def log_audit(action, details=''):
    """يُسجِّل عملية في سجل التدقيق"""
    try:
        username = session.get('username', 'system')
        execute('INSERT INTO audit_log (username, action, details) VALUES (?,?,?)',
                (username, action, details))
    except Exception:
        pass


def get_public_stats():
    """إحصائيات خفيفة للصفحات العامة مثل الصفحة الرئيسية وتسجيل الدخول"""
    latest_year = query('SELECT COALESCE(MAX(year), 0) AS v FROM institutions', one=True)['v']  # آخر سنة في البيانات
    latest_year = int(latest_year or 0)  # تحويل لعدد صحيح

    latest_totals = {'energy': 0.0, 'carbon': 0.0, 'renewable': 0.0}  # قيم افتراضية إذا لم تكن هناك بيانات
    if latest_year:  # إذا وُجدت بيانات
        row = query(
            'SELECT COALESCE(SUM(energy_consumption),0) AS total_energy, '
            'COALESCE(SUM(carbon_emissions),0) AS total_carbon, '
            'COALESCE(AVG(renewable_energy_percentage),0) AS avg_renewable '
            'FROM institutions WHERE year=?',  # جلب مجاميع الطاقة والكربون ومتوسط الطاقة المتجددة
            (latest_year,),
            one=True,
        )
        if row:  # إذا وُجدت بيانات
            latest_totals = {
                'energy': float(row['total_energy'] or 0),      # إجمالي استهلاك الطاقة
                'carbon': float(row['total_carbon'] or 0),      # إجمالي انبعاثات الكربون
                'renewable': float(row['avg_renewable'] or 0),  # متوسط الطاقة المتجددة
            }

    return {
        'total_institutions': int(query('SELECT COUNT(DISTINCT name) AS v FROM institutions', one=True)['v']),  # عدد المؤسسات الفريدة
        'years_count': int(query('SELECT COUNT(DISTINCT year) AS v FROM institutions', one=True)['v']),         # عدد السنوات المتاحة
        'indicator_count': 6,      # عدد المؤشرات البيئية (ثابت)
        'role_count': 3,           # عدد الأدوار (مسؤول / مستخدم / مشاهد)
        'latest_year': latest_year,                                # آخر سنة بيانات
        'latest_energy': round(latest_totals['energy'], 2),       # إجمالي استهلاك الطاقة
        'latest_carbon': round(latest_totals['carbon'], 2),       # إجمالي انبعاثات الكربون
        'latest_renewable': round(latest_totals['renewable'], 1), # متوسط الطاقة المتجددة
    }


# ── Rate Limiting بسيط (بدون مكتبات خارجية) ─────────────────────────────────
_rate_store = defaultdict(list)  # قاسوس يخزّن تواريخ محاولات كل IP
_RATE_MAX   = 10        # عدد المحاولات المسموح بها في النافذة
_RATE_WIN   = 60        # نافزة الوقت بالثواني

def _is_rate_limited(key):
    """ تتحقق إذا كان المفتاح (عادةً IP) تجاوز الحد المسموح به """
    now = time.time()  # الوقت الحالي
    _rate_store[key] = [t for t in _rate_store[key] if now - t < _RATE_WIN]  # احتفظ فقط بالمحاولات داخل النافذة
    if len(_rate_store[key]) >= _RATE_MAX:  # إذا وصل العدد للحد الأقصى
        return True   # موقوف – تجاوز الحد
    _rate_store[key].append(now)  # سجّل هذه المحاولة
    return False  # غير موقوف – لم يتجاوز الحد


# ── مصمم للتحقق من تسجيل الدخول ─────────────────────────────────────────────
def login_required(f):
    @wraps(f)  # احفظ اسم الدالة الأصلية
    def decorated(*args, **kwargs):
        if 'user_id' not in session:  # إذا لم يكن المستخدم مسجّل الدخول
            if request.path.startswith('/api/'):  # إذا كان طلباً API
                return jsonify({'error': 'انتهت الجلسة، يرجى تسجيل الدخول مجدداً'}), 401  # أرجع خطأ 401
            return redirect(url_for('login'))  # وإلا أعد توجيهه لصفحة الدخول
        return f(*args, **kwargs)  # نفّذ الدالة الأصلية
    return decorated  # أرجع الدالة المغلّفة


def role_required(*allowed_roles):
    """مصمم للتحقق من صلاحيات المستخدم (admin / user / viewer)"""
    def decorator(f):
        @wraps(f)  # احفظ اسم الدالة
        def decorated(*args, **kwargs):
            user_role = session.get('role', 'viewer')          # دور المستخدم الحالي
            if user_role not in allowed_roles:                 # إذا لم يكن له صلاحية
                if request.path.startswith('/api/'):           # إذا كان طلبا API
                    return jsonify({'error': 'ليس لديك صلاحية لهذا الإجراء'}), 403  # 403 Forbidden
                return render_template('error.html', code=403, message='ليس لديك صلاحية للوصول'), 403  # عرض صفحة خطأ
            return f(*args, **kwargs)   # نفذ الدالة الأصلية
        return decorated
    return decorator


# ══════════════════════════════════════════════════════════════════════════════
# صفحات HTML
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/')  # المسار الرئيسي للموقع
def index():
    public_stats = get_public_stats()                       # جلب الإحصائيات العامة
    if 'user_id' in session:                                # إذا كان المستخدم مسجل الدخول
        return redirect(url_for('dashboard'))               # أعد توجيهه للوحة التحكم
    return render_template('landing.html', public_stats=public_stats)  # عرض صفحة الترحيب


@app.route('/favicon.ico')  # أيقونة المتصفح
def favicon():
    # خدم أيقونة الموقع من مجلد static
    return send_from_directory(
        os.path.join(app.root_path, 'static'),  # مجلد الأيقونة
        'favicon.ico',
        mimetype='image/x-icon'
    )


@app.route('/login', methods=['GET', 'POST'])  # GET لعرض الصفحة POST لإرسال البيانات
def login():
    public_stats = get_public_stats()  # جلب الإحصائيات للصفحة
    if 'user_id' in session:          # إذا كان مسجل الدخول مسبقا
        return redirect(url_for('dashboard'))  # أعد توجيهه

    if request.method == 'POST':  # إذا أرسل المستخدم النموذج
        username = request.form.get('username', '').strip()  # اسم المستخدم من النموذج
        password = request.form.get('password', '')           # كلمة المرور من النموذج

        # Rate limiting على عمليات تسجيل الدخول
        client_ip = request.remote_addr or 'unknown'  # عنوان IP للزائر
        if _is_rate_limited(f'login:{client_ip}'):   # تحقق من تجاوز الحد
            flash('عدد المحاولات تجاوز الحد المسموح. حاول مرة أخرى بعد دقيقة.', 'error')  # رسالة تحذير
            return render_template('login.html', public_stats=public_stats)

        if not username or not password:  # إذا كان أحد الحقلين فارغا
            flash('يرجى إدخال اسم المستخدم وكلمة المرور', 'error')
            return render_template('login.html', public_stats=public_stats)

        user = query('SELECT id, username, password_hash, role FROM users WHERE username = ?', (username,), one=True)  # ابحث عن المستخدم

        if user and check_password_hash(user['password_hash'], password):  # إذا وجد وكلمة المرور صحيحة
            session['user_id'] = user['id']       # احفظ ID المستخدم في الجلسة
            session['username'] = user['username'] # احفظ اسمه
            session['role'] = user['role']         # احفظ دوره
            return redirect(url_for('dashboard'))  # أعد توجيهه للوحة التحكم

        flash('اسم المستخدم أو كلمة المرور غير صحيحة', 'error')  # كلمة السر غلط

    return render_template('login.html', public_stats=public_stats)  # عرض صفحة الدخول


@app.route('/logout')  # مسار تسجيل الخروج
def logout():
    session.clear()                          # مسح كل بيانات الجلسة
    return redirect(url_for('login'))        # إعادة التوجيه لصفحة تسجيل الدخول


@app.route('/dashboard')  # الصفحة الرئيسية للمستخدم
@login_required            # يجب تسجيل الدخول
def dashboard():
    return render_template('dashboard.html', username=session.get('username'), role=session.get('role', 'viewer'))  # عرض الصفحة مع اسم المستخدم ودوره


@app.route('/data-entry')  # صفحة إدخال البيانات
@login_required             # يجب تسجيل الدخول
@role_required('admin', 'user')  # يجب أن يكون المستخدم admin أو user
def data_entry():
    return render_template('data_entry.html', username=session.get('username'), role=session.get('role', 'viewer'))  # عرض صفحة إدخال البيانات


@app.route('/reports')  # صفحة التقارير
@login_required          # يجب تسجيل الدخول
def reports():
    return render_template('reports.html', username=session.get('username'), role=session.get('role', 'viewer'))  # عرض صفحة التقارير


@app.route('/green-score')  # مسار صفحة مؤشر الأداء البيئي
@login_required  # يجب تسجيل الدخول
def green_score_page():  # دالة عرض صفحة الدرجة البيئية
    return render_template('green_score.html', username=session.get('username'), role=session.get('role', 'viewer'))  # عرض صفحة مؤشر الأداء البيئي


@app.route('/predictions')  # مسار صفحة التنبؤات المستقبلية
@login_required  # يجب تسجيل الدخول
def predictions_page():  # دالة عرض صفحة التنبؤات
    return render_template('predictions.html', username=session.get('username'), role=session.get('role', 'viewer'))  # عرض صفحة التنبؤات


@app.route('/compare')  # مسار صفحة المقارنة بين مؤسستين
@login_required  # يجب تسجيل الدخول
def compare_page():  # دالة عرض صفحة المقارنة
    return render_template('compare.html', username=session.get('username'), role=session.get('role', 'viewer'))  # عرض صفحة المقارنة


# ══════════════════════════════════════════════════════════════════════════════
# REST API
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/dashboard-stats', methods=['GET'])  # API: إحصائيات لوحة التحكم
@login_required
def api_dashboard_stats():
    # استخدام آخر سنة موجودة في البيانات (وليس السنة الحالية)
    latest = query('SELECT COALESCE(MAX(year), 0) AS v FROM institutions', one=True)['v']  # آخر سنة في البيانات
    display_year = latest if latest else __import__('datetime').date.today().year  # إذا لم توجد بيانات استخدم السنة الحالية

    total_institutions = query('SELECT COUNT(DISTINCT name) AS v FROM institutions', one=True)['v']  # عدد المؤسسات الفريدة
    total_energy   = query('SELECT COALESCE(SUM(energy_consumption),0) AS v FROM institutions WHERE year=?',  (display_year,), one=True)['v']  # إجمالي استهلاك الطاقة
    total_carbon   = query('SELECT COALESCE(SUM(carbon_emissions),0) AS v FROM institutions WHERE year=?',    (display_year,), one=True)['v']  # إجمالي انبعاثات الكربون
    total_projects = query('SELECT COALESCE(SUM(green_projects),0) AS v FROM institutions WHERE year=?',      (display_year,), one=True)['v']  # إجمالي المشاريع الخضراء
    avg_renewable  = query('SELECT COALESCE(AVG(renewable_energy_percentage),0) AS v FROM institutions WHERE year=?', (display_year,), one=True)['v']  # متوسط نسبة الطاقة المتجددة
    avg_carbon     = query('SELECT COALESCE(AVG(carbon_emissions),0) AS v FROM institutions WHERE year=?', (display_year,), one=True)['v']  # متوسط انبعاثات الكربون
    total_water    = query('SELECT COALESCE(SUM(water_usage),0) AS v FROM institutions WHERE year=?', (display_year,), one=True)['v']  # إجمالي استهلاك المياه
    avg_waste      = query('SELECT COALESCE(AVG(waste_recycling_percentage),0) AS v FROM institutions WHERE year=?', (display_year,), one=True)['v']  # متوسط نسبة إعادة التدوير
    top_carbon_row = query('SELECT name, carbon_emissions FROM institutions WHERE year=? ORDER BY carbon_emissions DESC LIMIT 1', (display_year,), one=True)  # المؤسسة الأعلى انبعاثات
    best_renew_row = query('SELECT name, renewable_energy_percentage FROM institutions WHERE year=? ORDER BY renewable_energy_percentage DESC LIMIT 1', (display_year,), one=True)  # المؤسسة الأفضل في الطاقة المتجددة
    years_count    = query('SELECT COUNT(DISTINCT year) AS v FROM institutions', one=True)['v']  # عدد السنوات المتاحة
    total_records  = query('SELECT COUNT(*) AS v FROM institutions', one=True)['v']  # إجمالي عدد السجلات

    _threshold = get_carbon_threshold()  # جلب حد الكربون من الإعدادات
    alerts = []  # قائمة التنبيهات
    if float(total_carbon) > _threshold:  # إذا تجاوز إجمالي الكربون الحد المسموح
        alerts.append({
            'type': 'danger',
            'message': f'تحذير: إجمالي انبعاثات الكربون تجاوز الحد المسموح به ({_threshold} طن)'
        })

    return jsonify({
        'total_institutions': total_institutions,
        'total_energy':   round(float(total_energy),   2),
        'total_carbon':   round(float(total_carbon),   2),
        'total_projects': int(total_projects),
        'avg_renewable':  round(float(avg_renewable),  2),
        'avg_carbon':     round(float(avg_carbon),     2),
        'top_carbon_name':  top_carbon_row['name'] if top_carbon_row else '–',
        'top_carbon_value': round(float(top_carbon_row['carbon_emissions']), 1) if top_carbon_row else 0,
        'best_renew_name':  best_renew_row['name'] if best_renew_row else '–',
        'best_renew_value': round(float(best_renew_row['renewable_energy_percentage']), 1) if best_renew_row else 0,
        'years_count':    years_count,
        'total_records':  total_records,
        'total_water':    round(float(total_water), 2),
        'avg_waste':      round(float(avg_waste), 2),
        'alerts': alerts,
        'carbon_threshold': _threshold,
        'display_year': display_year,
    })


@app.route('/api/chart-data', methods=['GET'])  # API: بيانات الرسوم البيانية
@login_required
def api_chart_data():
    carbon_data   = query('SELECT name, carbon_emissions FROM institutions WHERE year=(SELECT MAX(year) FROM institutions) ORDER BY carbon_emissions DESC LIMIT 10')  # أعلى 10 مؤسسات في الانبعاثات
    energy_data   = query('SELECT name, energy_consumption FROM institutions WHERE year=(SELECT MAX(year) FROM institutions) ORDER BY energy_consumption DESC LIMIT 10')  # أعلى 10 مؤسسات في استهلاك الطاقة
    renewable_data = query('SELECT name, renewable_energy_percentage FROM institutions WHERE year=(SELECT MAX(year) FROM institutions) ORDER BY renewable_energy_percentage DESC LIMIT 10')  # أعلى 10 مؤسسات في الطاقة المتجددة
    trend_data    = query('SELECT year, SUM(carbon_emissions) AS total_carbon FROM institutions GROUP BY year ORDER BY year ASC')  # اتجاه الانبعاثات عبر السنين
    water_data    = query('SELECT name, water_usage FROM institutions WHERE year=(SELECT MAX(year) FROM institutions) ORDER BY water_usage DESC LIMIT 10')  # أعلى 10 مؤسسات في استهلاك المياه
    waste_data    = query('SELECT name, waste_recycling_percentage FROM institutions WHERE year=(SELECT MAX(year) FROM institutions) ORDER BY waste_recycling_percentage DESC LIMIT 10')  # أعلى 10 مؤسسات في إعادة التدوير

    return jsonify({
        'carbon':    {'labels': [r['name'] for r in carbon_data],    'values': [float(r['carbon_emissions'])           for r in carbon_data]},
        'energy':    {'labels': [r['name'] for r in energy_data],    'values': [float(r['energy_consumption'])          for r in energy_data]},
        'renewable': {'labels': [r['name'] for r in renewable_data], 'values': [float(r['renewable_energy_percentage']) for r in renewable_data]},
        'trend':     {'labels': [str(r['year']) for r in trend_data],'values': [float(r['total_carbon'])               for r in trend_data]},
        'water':     {'labels': [r['name'] for r in water_data],     'values': [float(r['water_usage'])                for r in water_data]},
        'waste':     {'labels': [r['name'] for r in waste_data],     'values': [float(r['waste_recycling_percentage']) for r in waste_data]},
    })


@app.route('/api/institutions', methods=['GET'])  # API: جلب قائمة المؤسسات مع دعم الفلترة والصفحات
@login_required
def api_get_institutions():
    year   = request.args.get('year')              # فلتر السنة (اختياري)
    name   = request.args.get('name', '').strip()  # فلتر اسم المؤسسة (اختياري)
    page   = max(int(request.args.get('page', 1)), 1)       # رقم الصفحة (لا يقل عن 1)
    limit  = min(int(request.args.get('limit', 20)), 100)   # عدد النتائج في الصفحة (حد أقصى 100)
    offset = (page - 1) * limit  # حساب عدد السجلات للتخطي

    conditions, params = [], []
    if year:  conditions.append('year = ?');        params.append(year)
    if name:  conditions.append('name LIKE ?');     params.append(f'%{name}%')
    where = 'WHERE ' + ' AND '.join(conditions) if conditions else ''

    total = query(f'SELECT COUNT(*) AS v FROM institutions {where}', params, one=True)['v']
    rows  = query(f'SELECT * FROM institutions {where} ORDER BY year DESC, name ASC LIMIT ? OFFSET ?', params + [limit, offset])

    return jsonify({
        'data':  [dict(r) for r in rows],
        'total': total,
        'page':  page,
        'limit': limit,
        'pages': (total + limit - 1) // limit,
    })


@app.route('/api/institutions', methods=['POST'])  # API: إضافة مؤسسة جديدة
@login_required
@role_required('admin', 'user')  # يحتاج صلاحية admin أو user
def api_add_institution():
    data = request.get_json(silent=True)  # قراءة بيانات JSON من جسم الطلب
    if not data:  # إذا لم توجد بيانات
        return jsonify({'error': 'بيانات غير صالحة'}), 400  # خطأ 400

    required = ['name','year','energy_consumption','renewable_energy_percentage','carbon_emissions','green_projects']
    for field in required:
        if field not in data or data[field] == '' or data[field] is None:
            return jsonify({'error': f'الحقل {field} مطلوب'}), 400

    try:
        name      = str(data['name']).strip()  # اسم المؤسسة
        year      = int(data['year'])  # السنة
        energy    = float(data['energy_consumption'])  # استهلاك الطاقة
        renewable = float(data['renewable_energy_percentage'])  # نسبة الطاقة المتجددة
        carbon    = float(data['carbon_emissions'])  # انبعاثات الكربون
        projects  = int(data['green_projects'])  # عدد المشاريع الخضراء
        water     = float(data.get('water_usage', 0))  # استهلاك المياه (اختياري)
        waste     = float(data.get('waste_recycling_percentage', 0))  # نسبة إعادة التدوير (اختياري)

        if not name:                      raise ValueError('اسم المؤسسة لا يمكن أن يكون فارغاً')
        if year < 2000 or year > 2100:    raise ValueError('السنة يجب أن تكون بين 2000 و2100')
        if energy < 0:                    raise ValueError('استهلاك الطاقة يجب أن يكون موجباً')
        if not 0 <= renewable <= 100:     raise ValueError('نسبة الطاقة المتجددة يجب أن تكون بين 0 و100')
        if carbon < 0:                    raise ValueError('انبعاثات الكربون يجب أن تكون موجبة')
        if projects < 0:                  raise ValueError('عدد المشاريع يجب أن يكون موجباً')
        if water < 0:                     raise ValueError('استهلاك المياه يجب أن يكون موجباً')
        if not 0 <= waste <= 100:         raise ValueError('نسبة إعادة تدوير النفايات يجب أن تكون بين 0 و100')
    except (ValueError, TypeError) as e:
        return jsonify({'error': str(e)}), 400

    # التحقق من التكرار
    duplicate = query('SELECT id FROM institutions WHERE name=? AND year=?', (name, year), one=True)
    if duplicate:
        return jsonify({'error': f'يوجد سجل بالفعل للمؤسسة "{name}" للسنة {year}. استخدم التعديل لتحديث البيانات.'}), 409

    cur = execute(
        'INSERT INTO institutions (name,year,energy_consumption,renewable_energy_percentage,carbon_emissions,green_projects,water_usage,waste_recycling_percentage) VALUES (?,?,?,?,?,?,?,?)',
        (name, year, energy, renewable, carbon, projects, water, waste)
    )
    institution = dict(query('SELECT * FROM institutions WHERE id=?', (cur.lastrowid,), one=True))

    # إرسال تنبيه بريدي إن تجاوزت الانبعاثات الحد
    _threshold = get_carbon_threshold()
    log_audit('إضافة مؤسسة', f'{name} ({year})')
    if carbon > _threshold:
        send_email_alert(
            subject=f'[Green Economy] تحذير: انبعاثات مرتفعة – {name}',
            body=f'المؤسسة: {name}\nالسنة: {year}\nانبعاثات الكربون: {carbon} طن\nتجاوزت الحد المسموح ({_threshold} طن)'
        )

    return jsonify({
        'message': 'تم إضافة البيانات بنجاح',
        'data': institution,
        'carbon_alert': carbon > _threshold,
    }), 201


@app.route('/api/institutions/<int:inst_id>', methods=['GET'])  # API: جلب بيانات مؤسسة واحدة
@login_required  # يجب تسجيل الدخول
def api_get_institution(inst_id):  # دالة جلب مؤسسة بالـ ID
    row = query('SELECT * FROM institutions WHERE id=?', (inst_id,), one=True)  # ابحث عن المؤسسة بالـ ID
    if not row:
        return jsonify({'error': 'السجل غير موجود'}), 404
    return jsonify(dict(row))


@app.route('/api/institutions/<int:inst_id>', methods=['PUT'])  # API: تعديل بيانات مؤسسة
@login_required  # يجب تسجيل الدخول
@role_required('admin', 'user')  # يحتاج صلاحية admin أو user
def api_update_institution(inst_id):  # دالة تعديل مؤسسة بالـ ID
    data = request.get_json(silent=True)  # قراءة البيانات من الطلب
    if not data:
        return jsonify({'error': 'بيانات غير صالحة'}), 400

    if not query('SELECT id FROM institutions WHERE id=?', (inst_id,), one=True):
        return jsonify({'error': 'السجل غير موجود'}), 404

    allowed = {'name': str, 'year': int, 'energy_consumption': float,
               'renewable_energy_percentage': float, 'carbon_emissions': float, 'green_projects': int,
               'water_usage': float, 'waste_recycling_percentage': float}
    fields, params = [], []
    for key, cast in allowed.items():
        if key in data:
            fields.append(f'{key} = ?')  # أضف الحقل لقائمة التحديث
            params.append(cast(data[key]))  # أضف القيمة بعد تحويل النوع

    if not fields:
        return jsonify({'error': 'لا توجد بيانات للتحديث'}), 400

    fields.append("updated_at = datetime('now')")
    params.append(inst_id)
    execute(f"UPDATE institutions SET {', '.join(fields)} WHERE id=?", params)
    log_audit('تعديل مؤسسة', f'ID={inst_id}')
    institution = dict(query('SELECT * FROM institutions WHERE id=?', (inst_id,), one=True))
    
    # التحقق من تجاوز حد الكربون عند التعديل
    _threshold = get_carbon_threshold()
    carbon = float(institution.get('carbon_emissions', 0))
    if carbon > _threshold:
        send_email_alert(
            subject=f'[Green Economy] تحذير: انبعاثات مرتفعة – {institution["name"]}',
            body=f'المؤسسة: {institution["name"]}\nالسنة: {institution["year"]}\nانبعاثات الكربون: {carbon} طن\nتجاوزت الحد المسموح ({_threshold} طن)'
        )
    
    return jsonify({'message': 'تم التحديث بنجاح', 'data': institution, 'carbon_alert': carbon > _threshold})


@app.route('/api/institutions/<int:inst_id>', methods=['DELETE'])  # API: حذف مؤسسة
@login_required
@role_required('admin')  # فقط المسؤول يستطيع الحذف
def api_delete_institution(inst_id):
    existing = query('SELECT id, name FROM institutions WHERE id=?', (inst_id,), one=True)  # ابحث عن المؤسسة
    if not existing:  # إذا لم توجد
        return jsonify({'error': 'السجل غير موجود'}), 404  # خطأ 404
    execute('DELETE FROM institutions WHERE id=?', (inst_id,))  # احذف المؤسسة
    log_audit('حذف مؤسسة', f'{existing["name"]} (ID={inst_id})')  # سجّل في سجل التدقيق
    return jsonify({'message': 'تم الحذف بنجاح'})


@app.route('/api/reports', methods=['GET'])  # API: جلب التقارير السنوية
@login_required  # يجب تسجيل الدخول
def api_reports():  # دالة التقارير السنوية
    yearly_summary = query("""
        SELECT year,
               COUNT(DISTINCT name)              AS institutions_count,
               SUM(energy_consumption)           AS total_energy,
               AVG(renewable_energy_percentage)  AS avg_renewable,
               SUM(carbon_emissions)             AS total_carbon,
               SUM(green_projects)               AS total_projects,
               SUM(water_usage)                  AS total_water,
               AVG(waste_recycling_percentage)   AS avg_waste
        FROM institutions
        GROUP BY year
        ORDER BY year DESC
    """)
    top_emitters  = query('SELECT name, year, carbon_emissions FROM institutions WHERE year=(SELECT MAX(year) FROM institutions) ORDER BY carbon_emissions DESC LIMIT 5')  # أعلى 5 مؤسسات في الانبعاثات
    top_renewable = query('SELECT name, year, renewable_energy_percentage FROM institutions WHERE year=(SELECT MAX(year) FROM institutions) ORDER BY renewable_energy_percentage DESC LIMIT 5')  # أفضل 5 مؤسسات في الطاقة المتجددة

    return jsonify({
        'yearly_summary': [dict(r) for r in yearly_summary],
        'top_emitters':   [dict(r) for r in top_emitters],
        'top_renewable':  [dict(r) for r in top_renewable],
    })


@app.route('/api/init-admin', methods=['POST'])  # API: إنشاء أول حساب مسؤول
def api_init_admin():  # دالة إنشاء حساب المسؤول الأولي
    data   = request.get_json(silent=True) or {}  # قراءة بيانات الطلب
    secret = data.get('setup_secret', '')  # مفتاح الإعداد السري
    if secret != os.environ.get('SETUP_SECRET', 'init-green-2024'):  # تحقق من صحة المفتاح
        return jsonify({'error': 'غير مصرح'}), 403

    username = data.get('username', 'admin')  # اسم المستخدم (الافتراضي admin)
    password = data.get('password', 'admin123')  # كلمة المرور (الافتراضي admin123)

    if query('SELECT id FROM users WHERE username=?', (username,), one=True):
        return jsonify({'message': 'المستخدم موجود بالفعل'})

    execute('INSERT INTO users (username, password_hash, role) VALUES (?,?,?)',
            (username, generate_password_hash(password), 'admin'))
    return jsonify({'message': f'تم إنشاء المستخدم {username} بنجاح'}), 201


# ══════════════════════════════════════════════════════════════════════════════
# API – Green Score (مؤشر الأداء البيئي المركّب)
# ══════════════════════════════════════════════════════════════════════════════

def calc_green_score(carbon, renewable_pct, green_projects, energy, water=0, waste_pct=0):
    """
    حساب درجة الأداء البيئي من 0 إلى 100:
    - انبعاثات الكربون (منخفض = أفضل)  : 30%
    - نسبة الطاقة المتجددة              : 25%
    - المشاريع الخضراء                  : 10%
    - كفاءة الطاقة (inversed)           : 10%
    - كفاءة استخدام المياه (منخفض = أفضل) : 15%
    - نسبة إعادة تدوير النفايات        : 10%
    """
    carbon_score    = max(0, 100 - (carbon / CARBON_ALERT_THRESHOLD) * 60)  # نقاط الكربون: قلّت الانبعاثات كلما ارتفعت
    renewable_score = min(100, renewable_pct)                                    # نقاط الطاقة المتجددة
    project_score   = min(100, green_projects * 5)                               # نقاط المشاريع الخضراء (5 نقاط لكل مشروع)
    energy_score    = max(0, 100 - (energy / 25000) * 100)                       # كفاءة الطاقة (استهلاك أقل = نقاط أعلى)
    water_score     = max(0, 100 - (water / 30000) * 100)                        # كفاءة المياه (استهلاك أقل = نقاط أعلى)
    waste_score     = min(100, waste_pct)                                         # نقاط إعادة التدوير

    score = (carbon_score * 0.30 + renewable_score * 0.25 +
             project_score * 0.10 + energy_score * 0.10 +
             water_score * 0.15 + waste_score * 0.10)  # الدرجة الإجمالية بالأوزان المحددة
    return round(min(100, max(0, score)), 1)  # اضمن أن الدرجة بين 0 و100


@app.route('/api/green-scores', methods=['GET'])  # API: حساب درجات الأداء البيئي
@login_required  # يجب تسجيل الدخول
def api_green_scores():  # دالة حساب الدرجات البيئية
    latest_year = query('SELECT COALESCE(MAX(year), 0) AS v FROM institutions', one=True)['v']  # آخر سنة في البيانات
    year = request.args.get('year') or str(latest_year)  # السنة المطلوبة أو آخر سنة
    rows = query(
        'SELECT id, name, year, carbon_emissions, renewable_energy_percentage, green_projects, energy_consumption, water_usage, waste_recycling_percentage '
        'FROM institutions WHERE year=? ORDER BY name',
        (year,)
    )
    result = []  # قائمة النتائج
    for r in rows:  # لكل مؤسسة في البيانات
        score = calc_green_score(
            float(r['carbon_emissions']),
            float(r['renewable_energy_percentage']),
            int(r['green_projects']),
            float(r['energy_consumption']),
            float(r['water_usage']),
            float(r['waste_recycling_percentage'])
        )
        grade = ('A+' if score >= 90 else 'A' if score >= 80 else 'B' if score >= 65
                 else 'C' if score >= 50 else 'D' if score >= 35 else 'F')  # التقدير حسب الدرجة
        result.append({**dict(r), 'green_score': score, 'grade': grade})  # أضف النتيجة للقائمة
    result.sort(key=lambda x: x['green_score'], reverse=True)  # رتّب تنازلياً حسب الدرجة
    return jsonify({'data': result, 'year': year})  # أرجع النتائج كـ JSON


# ══════════════════════════════════════════════════════════════════════════════
# API – AI Predictions (Linear Regression بدون مكتبات خارجية)
# ══════════════════════════════════════════════════════════════════════════════

def linear_regression(xs, ys):
    """ حساب معادلة الانحدار الخطي y = mx + b (بدون مكتبات خارجية) """
    n = len(xs)  # عدد نقاط البيانات
    if n < 2:   # نحتاج على الأقل نقطتين للتنبؤ
        return None, None
    sx, sy, sxy, sxx = sum(xs), sum(ys), sum(x*y for x,y in zip(xs,ys)), sum(x*x for x in xs)  # مجاميع للمعادلة
    denom = n * sxx - sx * sx  # مقام المعادلة
    if denom == 0:  # تجنب القسمة على صفر
        return None, None
    m = (n * sxy - sx * sy) / denom  # ميل الخط
    b = (sy - m * sx) / n            # ثابت الخط
    return m, b  # أرجع المعاملات


@app.route('/api/predictions', methods=['GET'])  # API: التنبؤات المستقبلية
@login_required  # يجب تسجيل الدخول
def api_predictions():  # دالة حساب التنبؤات
    name = request.args.get('name', '').strip()  # اسم المؤسسة (اختياري – فارغ = الكل)
    rows = query(
        'SELECT year, SUM(carbon_emissions) AS carbon, SUM(energy_consumption) AS energy, '
        'SUM(water_usage) AS water, AVG(waste_recycling_percentage) AS waste '
        'FROM institutions ' + ('WHERE name=?' if name else '') +
        ' GROUP BY year ORDER BY year',
        (name,) if name else ()
    )
    if len(rows) < 2:
        return jsonify({'error': 'بيانات غير كافية للتنبؤ (مطلوب سنتان على الأقل)'}), 400

    years  = [int(r['year'])         for r in rows]  # قائمة السنوات
    carbon = [float(r['carbon'])     for r in rows]  # قائمة انبعاثات الكربون
    energy = [float(r['energy'])     for r in rows]  # قائمة استهلاك الطاقة
    water  = [float(r['water'])      for r in rows]  # قائمة استهلاك المياه
    waste  = [float(r['waste'])      for r in rows]  # قائمة نسب إعادة التدوير

    mc, bc = linear_regression(years, carbon)  # حساب الانحدار الخطي للكربون
    me, be = linear_regression(years, energy)  # حساب الانحدار الخطي للطاقة
    mw, bw = linear_regression(years, water)  # حساب الانحدار الخطي للمياه
    mwa, bwa = linear_regression(years, waste)  # حساب الانحدار الخطي لإعادة التدوير

    future_years = [max(years) + 1, max(years) + 2, max(years) + 3]  # السنوات الـ 3 القادمة
    predictions = []  # قائمة التنبؤات
    for y in future_years:  # لكل سنة مستقبلية
        predictions.append({
            'year':   y,
            'carbon': round(mc * y + bc, 2) if mc is not None else None,
            'energy': round(me * y + be, 2) if me is not None else None,
            'water':  round(mw * y + bw, 2) if mw is not None else None,
            'waste':  round(max(0, min(100, mwa * y + bwa)), 1) if mwa is not None else None,
        })

    return jsonify({
        'historical': [{'year': r['year'], 'carbon': float(r['carbon']), 'energy': float(r['energy']),
                         'water': float(r['water']), 'waste': float(r['waste'])} for r in rows],
        'predictions': predictions,
        'institution': name or 'all',
    })


# ══════════════════════════════════════════════════════════════════════════════
# API – Compare Two Institutions
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/compare', methods=['GET'])  # API: مقارنة مؤسستين
@login_required  # يجب تسجيل الدخول
def api_compare():  # دالة المقارنة بين مؤسستين
    a = request.args.get('a', '').strip()  # اسم المؤسسة الأولى
    b = request.args.get('b', '').strip()  # اسم المؤسسة الثانية
    if not a or not b:
        return jsonify({'error': 'يجب تحديد مؤسستين للمقارنة'}), 400

    def get_stats(name):  # دالة داخلية لجلب إحصائيات مؤسسة
        rows = query(
            'SELECT year, carbon_emissions, energy_consumption, '
            'renewable_energy_percentage, green_projects, water_usage, waste_recycling_percentage FROM institutions '
            'WHERE name=? ORDER BY year',
            (name,)
        )
        if not rows:  # إذا لم توجد بيانات
            return None  # أرجع None
        latest = rows[-1]  # آخر سنة بيانات
        score = calc_green_score(
            float(latest['carbon_emissions']),
            float(latest['renewable_energy_percentage']),
            int(latest['green_projects']),
            float(latest['energy_consumption']),
            float(latest['water_usage']),
            float(latest['waste_recycling_percentage'])
        )
        return {
            'name': name,
            'records': [dict(r) for r in rows],
            'latest': dict(latest),
            'green_score': score,
            'grade': ('A+' if score >= 90 else 'A' if score >= 80 else 'B' if score >= 65
                      else 'C' if score >= 50 else 'D' if score >= 35 else 'F'),
        }

    stats_a = get_stats(a)  # إحصائيات المؤسسة الأولى
    stats_b = get_stats(b)  # إحصائيات المؤسسة الثانية

    if not stats_a:
        return jsonify({'error': f'لم يتم العثور على مؤسسة: {a}'}), 404
    if not stats_b:
        return jsonify({'error': f'لم يتم العثور على مؤسسة: {b}'}), 404

    return jsonify({'a': stats_a, 'b': stats_b})


# ══════════════════════════════════════════════════════════════════════════════
# API – Import CSV
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/import-csv', methods=['POST'])  # API: استيراد بيانات من ملف CSV أو Excel
@login_required  # يجب تسجيل الدخول
@role_required('admin', 'user')  # يحتاج صلاحية admin أو user
def api_import_csv():  # دالة استيراد البيانات
    import csv  # مكتبة قراءة ملفات CSV
    f = request.files.get('file')  # الملف المرفوع من المستخدم
    if not f:
        return jsonify({'error': 'يرجى رفع ملف'}), 400
    filename = f.filename.lower()  # اسم الملف بحروف صغيرة
    if not (filename.endswith('.csv') or filename.endswith('.xlsx')):
        return jsonify({'error': 'يرجى رفع ملف CSV أو Excel (.xlsx) صالح'}), 400

    required_cols = {'name', 'year', 'energy_consumption',
                     'renewable_energy_percentage', 'carbon_emissions', 'green_projects'}  # الأعمدة المطلوبة

    if filename.endswith('.xlsx'):
        try:
            import openpyxl
        except ImportError:
            return jsonify({'error': 'مكتبة openpyxl غير مثبتة. شغّل: pip install openpyxl'}), 500
        wb = openpyxl.load_workbook(io.BytesIO(f.read()))  # فتح ملف Excel
        ws = wb.active  # الورقة النشطة
        all_rows = list(ws.iter_rows(values_only=True))  # قراءة كل الصفوف
        if not all_rows:
            return jsonify({'error': 'الملف فارغ'}), 400
        header = [str(h).strip().lower() if h else '' for h in all_rows[0]]  # استخراج أسماء الأعمدة
        data_rows = [
            {col: (str(val).strip() if val is not None else '') for col, val in zip(header, row)}
            for row in all_rows[1:]
        ]
    else:
        content = f.read().decode('utf-8-sig')  # قراءة محتوى الملف بترميز UTF-8
        reader  = csv.DictReader(io.StringIO(content))  # قراءة CSV كقاموس لكل صف
        data_rows = [{k.strip().lower(): v.strip() for k, v in row.items()} for row in reader]

    inserted, skipped, errors = 0, 0, []  # عدّادات: مُدخَل، متخطّى، أخطاء
    for i, row in enumerate(data_rows, start=2):  # لكل صف من البيانات
        if not any(row.values()):
            continue
        missing = required_cols - set(row.keys())
        if missing:
            errors.append(f'سطر {i}: أعمدة مفقودة: {missing}')
            continue
        try:
            row_name = str(row['name']).strip()
            row_year = int(row['year'])
            if not row_name:
                errors.append(f'سطر {i}: اسم المؤسسة فارغ')
                continue
            # تحقق من التكرار
            dup = query('SELECT id FROM institutions WHERE name=? AND year=?', (row_name, row_year), one=True)
            if dup:
                skipped += 1
                continue
            execute(
                'INSERT INTO institutions (name,year,energy_consumption,renewable_energy_percentage,carbon_emissions,green_projects,water_usage,waste_recycling_percentage) VALUES (?,?,?,?,?,?,?,?)',
                (row_name, row_year, float(row['energy_consumption']),
                 float(row['renewable_energy_percentage']), float(row['carbon_emissions']),
                 int(row['green_projects']),
                 float(row.get('water_usage', 0)),
                 float(row.get('waste_recycling_percentage', 0)))
            )
            inserted += 1
        except Exception as e:
            errors.append(f'سطر {i}: {e}')

    log_audit('استيراد بيانات', f'{inserted} سجل من {f.filename}')
    msg = f'تم استيراد {inserted} سجل بنجاح'
    if skipped:  msg += f' (تم تخطي {skipped} سجل مكرر)'
    if errors:   msg += f' - {len(errors)} خطأ'
    return jsonify({'inserted': inserted, 'skipped': skipped, 'errors': errors, 'message': msg})


# ══════════════════════════════════════════════════════════════════════════════
# API – Export PDF
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/export-pdf', methods=['GET'])  # API: تصدير التقرير كـ PDF
@login_required  # يجب تسجيل الدخول
def api_export_pdf():  # دالة تصدير PDF
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return jsonify({'error': 'مكتبة reportlab غير مثبتة. شغّل: pip install reportlab'}), 500

    buf = io.BytesIO()  # مخزن مؤقت في الذاكرة لحفظ الـ PDF
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)  # إنشاء مستند PDF بحجم A4

    styles = getSampleStyleSheet()  # أنماط النصوص الجاهزة
    title_style = ParagraphStyle('title', parent=styles['Title'], fontSize=16, spaceAfter=10)  # نمط العنوان
    sub_style   = ParagraphStyle('sub',   parent=styles['Normal'], fontSize=9, textColor=colors.grey)  # نمط النص الفرعي

    elements = []  # قائمة عناصر الـ PDF
    elements.append(Paragraph('Green Economy Monitoring System', title_style))  # عنوان التقرير
    elements.append(Paragraph(f'Report generated: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}', sub_style))
    elements.append(Spacer(1, 0.4*cm))

    rows = query("""
        SELECT year, COUNT(DISTINCT name) institutions_count,
               ROUND(SUM(energy_consumption),1) total_energy,
               ROUND(AVG(renewable_energy_percentage),1) avg_renewable,
               ROUND(SUM(carbon_emissions),1) total_carbon,
               SUM(green_projects) total_projects
        FROM institutions GROUP BY year ORDER BY year DESC
    """)

    table_data = [['Year', 'Institutions', 'Energy (MWh)', 'Renewable %', 'Carbon (tons)', 'Green Projects']]
    for r in rows:
        table_data.append([str(r['year']), str(r['institutions_count']),
                           f"{r['total_energy']:,}", f"{r['avg_renewable']}%",
                           f"{r['total_carbon']:,}", str(r['total_projects'])])

    t = Table(table_data, repeatRows=1)  # إنشاء جدول مع تكرار الرأس في كل صفحة
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a6b3c')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTSIZE',   (0,0), (-1,0), 10),
        ('FONTSIZE',   (0,1), (-1,-1), 9),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f0f7f4')]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d4e8dc')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(t)

    doc.build(elements)  # بناء ملف الـ PDF
    buf.seek(0)  # إرجاع المؤشر لبداية الملف
    from flask import send_file
    return send_file(buf, mimetype='application/pdf',
                     as_attachment=True,
                     download_name=f'green-economy-{datetime.date.today()}.pdf')


# ══════════════════════════════════════════════════════════════════════════════
# API – Email Alert (إرسال تنبيه بريدي)
# ══════════════════════════════════════════════════════════════════════════════

def send_email_alert(subject, body):
    """ إرسال تنبيه بريدي عند تجاوز حد الكربون """
    if not MAIL_ENABLED or not MAIL_USERNAME or not MAIL_TO:  # تحقق من تفعيل البريد
        return False  # البريد غير مفعّل
    try:
        import smtplib                         # مكتبة إرسال البريد
        from email.mime.text import MIMEText  # لتنسيق رسالة البريد
        msg = MIMEText(body, 'plain', 'utf-8')  # أنشئ رسالة نصية
        msg['Subject'] = subject               # الموضوع
        msg['From']    = MAIL_USERNAME         # المُرسِل
        msg['To']      = MAIL_TO              # المُستقبِل
        with smtplib.SMTP(MAIL_SERVER, MAIL_PORT) as smtp:  # اتصل بالخادم
            smtp.starttls()                   # تشفير TLS
            smtp.login(MAIL_USERNAME, MAIL_PASSWORD)  # تسجيل الدخول
            smtp.send_message(msg)             # إرسال الرسالة
        return True   # نجح الإرسال
    except Exception:
        return False  # فشل الإرسال (لا تتوقف التطبيق)


# ══════════════════════════════════════════════════════════════════════════════
# API – Get Institution Names (للبحث والمقارنة)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/institution-names', methods=['GET'])  # API: جلب أسماء المؤسسات
@login_required  # يجب تسجيل الدخول
def api_institution_names():  # دالة جلب أسماء المؤسسات الفريدة
    rows = query('SELECT DISTINCT name FROM institutions ORDER BY name')  # جلب الأسماء مرتبة أبجدياً
    return jsonify([r['name'] for r in rows])  # أرجع قائمة الأسماء


@app.route('/api/years', methods=['GET'])  # API: جلب السنوات المتاحة
@login_required  # يجب تسجيل الدخول
def api_years():  # دالة جلب قائمة السنوات
    rows = query('SELECT DISTINCT year FROM institutions ORDER BY year DESC')  # جلب السنوات تنازلياً
    return jsonify([r['year'] for r in rows])  # أرجع قائمة السنوات


# ══════════════════════════════════════════════════════════════════════════════
# API – Export Excel
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/export-excel', methods=['GET'])  # API: تصدير التقرير كـ Excel
@login_required  # يجب تسجيل الدخول
def api_export_excel():  # دالة تصدير Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error': 'مكتبة openpyxl غير مثبتة. شغّل: pip install openpyxl'}), 500

    wb = openpyxl.Workbook()  # إنشاء ملف Excel جديد
    ws = wb.active  # الورقة النشطة
    ws.title = 'Green Economy Report'  # اسم الورقة
    ws.sheet_view.rightToLeft = True  # اتجاه من اليمين لليسار (للعربي)

    header_font  = Font(name='Arial', bold=True, color='FFFFFF', size=11)  # خط عنوان الجدول (أبيض عريض)
    header_fill  = PatternFill(fill_type='solid', fgColor='1a6b3c')  # خلفية العنوان (أخضر داكن)
    header_align = Alignment(horizontal='center', vertical='center')  # محاذاة العنوان للمنتصف
    alt_fill     = PatternFill(fill_type='solid', fgColor='f0f7f4')  # خلفية الصفوف المتناوبة
    thin_side    = Side(border_style='thin', color='d4e8dc')  # حدود رفيعة
    cell_border  = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)  # حدود كاملة للخلية

    headers    = ['السنة', 'عدد المؤسسات', 'إجمالي الطاقة (MWh)',
                  'متوسط الطاقة المتجددة (%)', 'إجمالي انبعاثات الكربون (طن)', 'إجمالي المشاريع الخضراء',
                  'إجمالي المياه (م³)', 'متوسط التدوير (%)']
    col_widths = [10, 18, 24, 28, 34, 28, 24, 22]

    for i, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = cell_border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28

    rows = query("""
        SELECT year,
               COUNT(DISTINCT name)                   AS institutions_count,
               ROUND(SUM(energy_consumption),1)       AS total_energy,
               ROUND(AVG(renewable_energy_percentage),1) AS avg_renewable,
               ROUND(SUM(carbon_emissions),1)         AS total_carbon,
               SUM(green_projects)                    AS total_projects,
               ROUND(SUM(water_usage),1)              AS total_water,
               ROUND(AVG(waste_recycling_percentage),1) AS avg_waste
        FROM institutions GROUP BY year ORDER BY year DESC
    """)

    for row_idx, r in enumerate(rows, start=2):  # لكل صف من البيانات
        fill   = alt_fill if row_idx % 2 == 0 else None  # تلوين متناوب للصفوف
        values = [r['year'], r['institutions_count'], r['total_energy'],
                  r['avg_renewable'], r['total_carbon'], r['total_projects'],
                  r['total_water'], r['avg_waste']]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)  # اكتب القيمة في الخلية
            cell.alignment = Alignment(horizontal='center')  # محاذاة للمنتصف
            cell.border    = cell_border  # أضف الحدود
            if fill:  # إذا كان الصف زوجي
                cell.fill = fill  # لوّن الخلفية

    buf = io.BytesIO()  # مخزن مؤقت في الذاكرة
    wb.save(buf)  # حفظ ملف Excel في المخزن
    buf.seek(0)  # إرجاع المؤشر لبداية الملف
    from flask import send_file  # استيراد دالة إرسال الملفات
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # نوع ملف Excel
        as_attachment=True,  # تحميل كملف مرفق
        download_name=f'green-economy-{datetime.date.today()}.xlsx'  # اسم ملف التحميل
    )


# ══════════════════════════════════════════════════════════════════════════════
# API – User Management
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/users', methods=['GET'])  # API: جلب قائمة المستخدمين
@login_required  # يجب تسجيل الدخول
@role_required('admin')  # فقط المسؤول
def api_get_users():  # دالة جلب المستخدمين
    rows = query('SELECT id, username, role, created_at FROM users ORDER BY created_at ASC')  # جلب كل المستخدمين مرتبين بتاريخ الإنشاء
    return jsonify([dict(r) for r in rows])  # أرجع القائمة كـ JSON


@app.route('/api/users', methods=['POST'])  # API: إضافة مستخدم جديد
@login_required  # يجب تسجيل الدخول
@role_required('admin')  # فقط المسؤول
def api_add_user():  # دالة إضافة مستخدم
    data     = request.get_json(silent=True) or {}  # قراءة البيانات من الطلب
    username = str(data.get('username', '')).strip()  # اسم المستخدم
    password = str(data.get('password', '')).strip()  # كلمة المرور
    role     = str(data.get('role', 'viewer')).strip()  # الصلاحية (الافتراضي: مشاهد)
    if role not in ('admin', 'user', 'viewer'):
        return jsonify({'error': 'الصلاحية يجب أن تكون admin أو user أو viewer'}), 400
    if not username or not password:
        return jsonify({'error': 'اسم المستخدم وكلمة المرور مطلوبان'}), 400
    if len(username) < 3 or len(username) > 50:
        return jsonify({'error': 'اسم المستخدم يجب أن يكون بين 3 و50 حرفاً'}), 400
    if len(password) < 6:
        return jsonify({'error': 'كلمة المرور يجب أن تكون 6 أحرف على الأقل'}), 400
    if query('SELECT id FROM users WHERE username=?', (username,), one=True):
        return jsonify({'error': 'اسم المستخدم موجود بالفعل'}), 409
    execute('INSERT INTO users (username, password_hash, role) VALUES (?,?,?)',
            (username, generate_password_hash(password), role))
    log_audit('إضافة مستخدم', f'{username} ({role})')
    return jsonify({'message': f'تم إنشاء المستخدم {username} بنجاح'}), 201


@app.route('/api/users/<int:user_id>', methods=['DELETE'])  # API: حذف مستخدم
@login_required  # يجب تسجيل الدخول
@role_required('admin')  # فقط المسؤول
def api_delete_user(user_id):  # دالة حذف مستخدم بالـ ID
    if session.get('user_id') == user_id:
        return jsonify({'error': 'لا يمكنك حذف حسابك الخاص'}), 403
    user = query('SELECT username FROM users WHERE id=?', (user_id,), one=True)
    if not user:
        return jsonify({'error': 'المستخدم غير موجود'}), 404
    execute('DELETE FROM users WHERE id=?', (user_id,))
    log_audit('حذف مستخدم', user['username'])
    return jsonify({'message': 'تم حذف المستخدم بنجاح'})


# ══════════════════════════════════════════════════════════════════════════════
# API – Audit Log
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/audit-log', methods=['GET'])  # API: جلب سجل التدقيق
@login_required  # يجب تسجيل الدخول
def api_audit_log_view():  # دالة عرض سجل التدقيق
    limit = min(int(request.args.get('limit', 50)), 200)  # الحد الأقصى للنتائج (200 كحد أقصى)
    rows  = query('SELECT * FROM audit_log ORDER BY created_at DESC LIMIT ?', (limit,))  # جلب آخر العمليات
    return jsonify([dict(r) for r in rows])  # أرجع السجل كـ JSON


# ══════════════════════════════════════════════════════════════════════════════
# صفحة الإعدادات
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/settings')  # مسار صفحة الإعدادات
@login_required  # يجب تسجيل الدخول
@role_required('admin')  # فقط المسؤول
def settings_page():  # دالة عرض صفحة الإعدادات
    return render_template('settings.html', username=session.get('username'), role=session.get('role', 'viewer'))  # عرض صفحة الإعدادات


# ══════════════════════════════════════════════════════════════════════════════
# صفحة الإشعارات والتنبيهات
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/notifications')  # مسار صفحة الإشعارات
@login_required  # يجب تسجيل الدخول
def notifications_page():  # دالة عرض صفحة الإشعارات
    return render_template('notifications.html', username=session.get('username'), role=session.get('role', 'viewer'))  # عرض صفحة الإشعارات


@app.route('/api/notifications', methods=['GET'])  # API: جلب الإشعارات والتنبيهات
@login_required  # يجب تسجيل الدخول
def api_notifications():  # دالة جلب الإشعارات
    threshold = get_carbon_threshold()  # جلب حد الكربون
    latest_year = query('SELECT COALESCE(MAX(year), 0) AS v FROM institutions', one=True)['v']  # آخر سنة في البيانات

    # تنبيهات: مؤسسات تجاوزت حد الكربون
    high_carbon = query(
        'SELECT name, year, carbon_emissions, renewable_energy_percentage '
        'FROM institutions WHERE carbon_emissions > ? ORDER BY carbon_emissions DESC LIMIT 15',
        (threshold,)
    )

    # أفضل 5 مؤسسات من حيث الطاقة المتجددة
    top_renewable = query(
        'SELECT name, renewable_energy_percentage FROM institutions '
        'WHERE year=? ORDER BY renewable_energy_percentage DESC LIMIT 5',
        (latest_year,)
    )

    # آخر نشاطات من سجل التدقيق
    recent_activity = query(
        'SELECT action, details, username, created_at FROM audit_log '
        'ORDER BY created_at DESC LIMIT 30'
    )

    # إحصاءات سريعة
    total_recs = query('SELECT COUNT(*) AS v FROM institutions', one=True)['v']  # إجمالي عدد السجلات
    total_high = len(high_carbon)  # عدد المؤسسات المتجاوزة للحد

    alerts = []  # قائمة التنبيهات
    for r in high_carbon:  # لكل مؤسسة تجاوزت الحد
        excess_pct = round((float(r['carbon_emissions']) - threshold) / threshold * 100, 1)  # نسبة التجاوز المئوية
        alerts.append({
            'type': 'danger' if excess_pct > 30 else 'warning',
            'name': r['name'],
            'year': r['year'],
            'carbon': round(float(r['carbon_emissions']), 1),
            'threshold': threshold,
            'excess_pct': excess_pct,
            'renewable': round(float(r['renewable_energy_percentage']), 1),
        })

    return jsonify({
        'alerts': alerts,
        'activities': [dict(r) for r in recent_activity],
        'top_renewable': [dict(r) for r in top_renewable],
        'carbon_threshold': threshold,
        'high_carbon_count': total_high,
        'total_records': total_recs,
        'latest_year': latest_year,
    })


@app.route('/api/settings', methods=['GET', 'POST'])  # API: قراءة وحفظ الإعدادات
@login_required  # يجب تسجيل الدخول
@role_required('admin')  # فقط المسؤول
def api_settings():  # دالة إدارة الإعدادات
    if request.method == 'GET':
        rows = query(f'SELECT {SETTINGS_KEY_COL} AS key, value FROM settings')  # جلب كل الإعدادات
        data = {r['key']: r['value'] for r in rows}  # تحويل لقاموس
        # معلومات النظام (للقراءة فقط)
        import sys, flask as _flask
        data['_sys_flask_version']       = _flask.__version__
        data['_sys_python_version']      = sys.version.split()[0]
        data['_sys_db_type']             = DB_TYPE_LABEL
        total = query('SELECT COUNT(*) AS n FROM institutions', one=True)
        data['_sys_total_institutions']  = str(total['n']) if total else '0'
        if USE_SQLITE:
            try:
                size_bytes = os.path.getsize(DB_PATH)
                data['_sys_db_size'] = f'{size_bytes / 1024:.1f} KB'
            except Exception:
                data['_sys_db_size'] = '–'
        else:
            data['_sys_db_size'] = '–'
        return jsonify(data)

    # POST – حفظ الإعدادات
    payload = request.get_json(silent=True) or {}  # قراءة الإعدادات المُرسلة
    for key, value in payload.items():  # لكل إعداد
        upsert_setting(key, str(value))
    return jsonify({'message': 'تم حفظ الإعدادات بنجاح'})


@app.route('/api/change-password', methods=['POST'])  # API: تغيير كلمة المرور
@login_required  # يجب تسجيل الدخول
def api_change_password():  # دالة تغيير كلمة المرور
    data = request.get_json(silent=True) or {}  # قراءة البيانات
    current_pwd = data.get('current_password', '').strip()  # كلمة المرور الحالية
    new_pwd     = data.get('new_password', '').strip()  # كلمة المرور الجديدة
    if not current_pwd or not new_pwd:
        return jsonify({'error': 'يرجى تعبئة جميع الحقول'}), 400
    if len(new_pwd) < 6:
        return jsonify({'error': 'كلمة المرور يجب أن تكون 6 أحرف على الأقل'}), 400
    username = session.get('username')  # اسم المستخدم الحالي من الجلسة
    user = query('SELECT * FROM users WHERE username=?', (username,), one=True)  # جلب بيانات المستخدم من قاعدة البيانات
    if not user or not check_password_hash(user['password_hash'], current_pwd):
        return jsonify({'error': 'كلمة المرور الحالية غير صحيحة'}), 401
    execute('UPDATE users SET password_hash=? WHERE username=?',
            (generate_password_hash(new_pwd), username))
    return jsonify({'message': 'تم تغيير كلمة المرور بنجاح'})


# ══════════════════════════════════════════════════════════════════════════════
# API – Backup & Data Management
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/backup', methods=['GET'])  # API: تحميل نسخة احتياطية
@login_required  # يجب تسجيل الدخول
@role_required('admin')  # فقط المسؤول
def api_backup():  # دالة إنشاء نسخة احتياطية
    if not USE_SQLITE:
        return jsonify({'error': 'النسخ الاحتياطي مدعوم لSQLite فقط'}), 400
    if not os.path.exists(DB_PATH):
        return jsonify({'error': 'ملف قاعدة البيانات غير موجود'}), 404
    backup_dir = os.path.join(os.path.dirname(DB_PATH), 'backups')  # مسار مجلد النسخ الاحتياطية
    os.makedirs(backup_dir, exist_ok=True)  # أنشئ المجلد إن لم يكن موجوداً
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')  # وسم الوقت الحالي
    backup_name = f'green_economy_backup_{timestamp}.db'  # اسم ملف النسخة
    backup_path = os.path.join(backup_dir, backup_name)  # المسار الكامل للنسخة
    shutil.copy2(DB_PATH, backup_path)  # نسخ قاعدة البيانات
    from flask import send_file  # استيراد دالة إرسال الملف
    return send_file(  # أرسل ملف النسخة للتحميل
        backup_path,
        mimetype='application/x-sqlite3',
        as_attachment=True,
        download_name=backup_name
    )


@app.route('/api/delete-all-data', methods=['DELETE'])  # API: حذف جميع البيانات
@login_required  # يجب تسجيل الدخول
@role_required('admin')  # فقط المسؤول
def api_delete_all_data():  # دالة حذف جميع البيانات
    confirm = request.args.get('confirm', '')  # معامل التأكيد
    if confirm != 'yes':
        return jsonify({'error': 'يجب تأكيد عملية الحذف'}), 400
    execute('DELETE FROM institutions')  # حذف كل بيانات المؤسسات
    execute('DELETE FROM audit_log')  # حذف كل سجلات التدقيق
    return jsonify({'message': 'تم حذف جميع البيانات بنجاح'})


@app.route('/api/restore-backup', methods=['POST'])  # API: استعادة نسخة احتياطية
@login_required  # يجب تسجيل الدخول
@role_required('admin')  # فقط المسؤول
def api_restore_backup():  # دالة استعادة النسخة الاحتياطية
    if not USE_SQLITE:
        return jsonify({'error': 'الاستعادة مدعومة لSQLite فقط'}), 400
    if 'file' not in request.files:
        return jsonify({'error': 'لم يتم إرسال ملف'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'error': 'لم يتم اختيار ملف'}), 400
    if not f.filename.endswith('.db'):
        return jsonify({'error': 'يجب أن يكون الملف بصيغة .db'}), 400

    # حفظ نسخة احتياطية من القاعدة الحالية قبل الاستعادة
    backup_dir = os.path.join(os.path.dirname(DB_PATH), 'backups')
    os.makedirs(backup_dir, exist_ok=True)
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    pre_restore = os.path.join(backup_dir, f'pre_restore_{timestamp}.db')
    if os.path.exists(DB_PATH):
        shutil.copy2(DB_PATH, pre_restore)

    # التحقق من صلاحية الملف المرفوع
    try:
        import tempfile
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.db')
        f.save(tmp.name)
        tmp.close()
        test_db = sqlite3.connect(tmp.name)
        test_db.execute('SELECT name FROM sqlite_master WHERE type="table"')
        tables = [r[0] for r in test_db.execute('SELECT name FROM sqlite_master WHERE type="table"').fetchall()]
        test_db.close()
        if 'institutions' not in tables:
            os.unlink(tmp.name)
            return jsonify({'error': 'الملف لا يحتوي على جدول institutions – نسخة غير صالحة'}), 400
        # استبدال قاعدة البيانات
        shutil.move(tmp.name, DB_PATH)
    except sqlite3.DatabaseError:
        if os.path.exists(tmp.name):
            os.unlink(tmp.name)
        return jsonify({'error': 'الملف ليس قاعدة بيانات SQLite صالحة'}), 400

    return jsonify({'message': 'تم استعادة النسخة الاحتياطية بنجاح – يُنصح بإعادة تشغيل التطبيق'})


# ══════════════════════════════════════════════════════════════════════════════
# معالجات الأخطاء
# ══════════════════════════════════════════════════════════════════════════════

@app.errorhandler(404)  # معالج خطأ 404 - الصفحة غير موجودة
def not_found(e):
    if request.path.startswith('/api/'):  # إذا كان طلباً API
        return jsonify({'error': 'المسار غير موجود'}), 404  # أرجع JSON
    return render_template('error.html', code=404, message='الصفحة غير موجودة'), 404  # عرض صفحة خطأ


@app.errorhandler(500)  # معالج خطأ 500 - خطأ داخلي في الخادم
def server_error(e):
    if request.path.startswith('/api/'):  # إذا كان طلباً API
        return jsonify({'error': 'خطأ داخلي في الخادم'}), 500  # أرجع JSON
    return render_template('error.html', code=500, message='خطأ داخلي في الخادم'), 500  # عرض صفحة خطأ


# ══════════════════════════════════════════════════════════════════════════════
# تهيئة قاعدة البيانات — تعمل دايمًا (سواء تشغيل مباشر أو عن طريق سيرفر
# خارجي زي Vercel/Gunicorn). قبل كده كانت جوه if __name__=='__main__' بس،
# فكانت بتتجاهَل تمامًا لما Vercel بيستورد الملف كـ module، فتطلع 500 error
# لأن الجداول مش بتتعمل خالص.
# ══════════════════════════════════════════════════════════════════════════════
try:
    init_db()  # أنشئ قاعدة البيانات والجداول إن لم تكن موجودة (آمن التكرار)
    with app.app_context():
        user = query('SELECT id FROM users WHERE username=?', ('admin',), one=True)  # تحقق من وجود حساب admin
        if not user:  # إذا لم يكن موجوداً
            execute('INSERT INTO users (username, password_hash, role) VALUES (?,?,?)',
                    ('admin', generate_password_hash('admin123'), 'admin'))  # أنشئ حساب admin بكلمة مرور 'admin123'
            print('[+] Admin account created: admin / admin123')
            print('[!] تحذير أمني: غيّر كلمة المرور الافتراضية فور تسجيل الدخول الأول!')
except Exception as e:
    print(f'\n[ERROR] فشلت تهيئة قاعدة البيانات: {e}')  # لو فشلت التهيئة، اطبع الخطأ بس متوقفش السيرفر


if __name__ == '__main__':  # يُنفَّذ فقط عند تشغيل الملف مباشرةً على جهازك
    import threading  # لتشغيل المتصفح في خيط منفصل
    import webbrowser  # لفتح المتصفح تلقائياً

    # ── التأكد من أن المجلد الحالي هو مجلد المشروع ──
    os.chdir(os.path.dirname(os.path.abspath(__file__)))  # انتقل لمجلد المشروع

    threading.Timer(1.5, lambda: webbrowser.open('http://127.0.0.1:5000')).start()  # افتح المتصفح بعد 1.5 ثانية
    app.run(debug=False, host='0.0.0.0', port=5000)  # شغّل الخادم على بورت 5000
